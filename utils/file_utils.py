# -*- coding: utf-8 -*-
"""
Утиліти для роботи з файлами.
"""

import os
import json
import csv
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional
from datetime import datetime

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


def ensure_directory_exists(directory_path: str) -> None:
    """
    Створює директорію, якщо вона не існує.

    Args:
        directory_path: Шлях до директорії
    """
    Path(directory_path).mkdir(parents=True, exist_ok=True)


def save_json_to_file(data: Any, file_path: str, ensure_ascii: bool = False, indent: int = 2) -> None:
    """
    Зберігає дані у JSON файл з кодуванням UTF-8.

    Args:
        data: Дані для збереження
        file_path: Шлях до файлу
        ensure_ascii: Чи екранувати не-ASCII символи
        indent: Відступ для форматування JSON
    """
    ensure_directory_exists(os.path.dirname(file_path) if os.path.dirname(file_path) else '.')
    
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=ensure_ascii, indent=indent, default=str)


def generate_json_filename(prefix: str = 'prozorro_real_estate', extension: str = 'json') -> str:
    """
    Генерує ім'я JSON файлу з поточною датою та часом.

    Args:
        prefix: Префікс імені файлу
        extension: Розширення файлу

    Returns:
        str: Ім'я файлу у форматі prefix_YYYY-MM-DD_HH-MM-SS.extension
    """
    now = datetime.now()
    timestamp = now.strftime('%Y-%m-%d_%H-%M-%S')
    return f'{prefix}_{timestamp}.{extension}'


def generate_auction_filename(
    prefix: str = 'prozorro_real_estate_auctions',
    extension: str = 'xlsx',
    user_id: Optional[int] = None,
    days: Optional[int] = None
) -> str:
    """
    Генерує ім'я файлу для аукціонів з датою, часом, ідентифікатором користувача та кількістю днів.

    Args:
        prefix: Префікс імені файлу
        extension: Розширення файлу
        user_id: Ідентифікатор користувача (опціонально)
        days: Кількість днів виборки (опціонально)

    Returns:
        str: Ім'я файлу у форматі prefix_YYYY-MM-DD_HH-MM-SS[_userID][_daysD].extension
    """
    now = datetime.now()
    timestamp = now.strftime('%Y-%m-%d_%H-%M-%S')
    
    parts = [prefix, timestamp]
    
    if user_id is not None:
        parts.append(str(user_id))
    
    if days is not None:
        parts.append(f'{days}D')
    
    filename = '_'.join(parts)
    return f'{filename}.{extension}'


def save_csv_to_file(data: List[Dict[str, Any]], file_path: str, fieldnames: List[str]) -> None:
    """
    Зберігає дані у CSV файл з кодуванням UTF-8 з BOM для кращої сумісності з Excel.
    Використовує крапку з комою (;) як роздільник та правильно екранує значення.

    Args:
        data: Список словників з даними для збереження
        file_path: Шлях до файлу
        fieldnames: Список назв колонок
    """
    ensure_directory_exists(os.path.dirname(file_path) if os.path.dirname(file_path) else '.')
    
    # Використовуємо UTF-8 з BOM для кращої сумісності з Excel та іншими програмами Windows
    with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
        # Використовуємо крапку з комою як роздільник
        # quoting=csv.QUOTE_MINIMAL - екранує значення, які містять роздільник, лапки або переноси рядків
        writer = csv.DictWriter(
            f, 
            fieldnames=fieldnames, 
            delimiter=';',
            quotechar='"',
            quoting=csv.QUOTE_MINIMAL,
            extrasaction='ignore'
        )
        writer.writeheader()
        writer.writerows(data)
    
    # Підтвердження збереження
    if os.path.exists(file_path):
        file_size = os.path.getsize(file_path)
        print(f"✓ CSV файл збережено: {file_path} ({file_size} байт)")
    else:
        print(f"⚠ Помилка: файл не знайдено після збереження: {file_path}")


def generate_excel_in_memory(data: List[Dict[str, Any]], fieldnames: List[str], column_headers: Optional[Dict[str, str]] = None) -> BytesIO:
    """
    Генерує Excel файл в пам'яті (BytesIO) з кодуванням UTF-8 та покращеним форматуванням.
    
    Args:
        data: Список словників з даними для збереження
        fieldnames: Список назв колонок (ключі)
        column_headers: Словник з українськими назвами колонок (ключ -> назва)
        
    Returns:
        BytesIO: Об'єкт з Excel файлом в пам'яті
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("Для збереження в Excel потрібно встановити pandas та openpyxl: pip install pandas openpyxl")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Для форматування Excel потрібно встановити openpyxl: pip install openpyxl")
    
    # Створюємо DataFrame з даних
    df = pd.DataFrame(data, columns=fieldnames)
    
    # Створюємо Workbook в пам'яті
    wb = Workbook()
    ws = wb.active
    
    # Заповнюємо дані
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # Стилі для шапки
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Стилі для комірок
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Форматуємо шапку
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx)
        # Використовуємо українську назву, якщо вказано
        if column_headers and fieldname in column_headers:
            cell.value = column_headers[fieldname]
        else:
            cell.value = fieldname
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Закріплюємо шапку
    ws.freeze_panes = 'A2'
    
    # Знаходимо індекс колонки з посиланнями
    url_column_idx = None
    if 'auction_url' in fieldnames:
        url_column_idx = fieldnames.index('auction_url') + 1
    
    # Стиль для гіперпосилань
    hyperlink_font = Font(underline="single", color="0563C1")
    # Стиль для жирного шрифту
    bold_font = Font(bold=True)
    
    # Форматуємо комірки з даними
    for row_idx in range(2, len(data) + 2):
        row_data = data[row_idx - 2]  # Індекс в масиві data
        # Перевіряємо, чи є додатковий класифікатор 03.07
        has_bold = row_data.get('_has_additional_classification_03_07', False)
        
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_alignment
            
            # Якщо це колонка з посиланнями, створюємо гіперпосилання
            if col_idx == url_column_idx and cell.value:
                url = str(cell.value).strip()
                if url and (url.startswith('http://') or url.startswith('https://')):
                    cell.hyperlink = url
                    cell.value = "Посилання"
                    # Комбінуємо стиль гіперпосилання з жирним, якщо потрібно
                    if has_bold:
                        cell.font = Font(underline="single", color="0563C1", bold=True)
                    else:
                        cell.font = hyperlink_font
                    cell.border = border_style
                elif cell.value and str(cell.value).strip():
                    if has_bold:
                        cell.font = bold_font
                    cell.border = border_style
                else:
                    cell.border = Border()
            # Додаємо сітку тільки якщо є дані
            elif cell.value and str(cell.value).strip():
                if has_bold:
                    cell.font = bold_font
                cell.border = border_style
            else:
                # Прибираємо сітку для порожніх комірок
                cell.border = Border()
    
    # Автоматично підганяємо ширину колонок
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Перевіряємо довжину в шапці
        cell = ws.cell(row=1, column=col_idx)
        if cell.value:
            max_length = len(str(cell.value))
        
        # Перевіряємо довжину в даних
        for row_idx in range(2, len(data) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_value = str(cell.value)
                # Для багаторядкових значень беремо найдовший рядок
                if '\n' in cell_value:
                    max_line_length = max(len(line) for line in cell_value.split('\n'))
                    max_length = max(max_length, max_line_length)
                else:
                    max_length = max(max_length, len(cell_value))
        
        # Встановлюємо ширину з невеликим запасом
        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символів
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Зберігаємо в BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def save_excel_to_file(data: List[Dict[str, Any]], file_path: str, fieldnames: List[str], column_headers: Optional[Dict[str, str]] = None) -> None:
    """
    Зберігає дані у Excel файл (.xlsx) з кодуванням UTF-8 та покращеним форматуванням.
    
    Args:
        data: Список словників з даними для збереження
        file_path: Шлях до файлу
        fieldnames: Список назв колонок (ключі)
        column_headers: Словник з українськими назвами колонок (ключ -> назва)
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("Для збереження в Excel потрібно встановити pandas та openpyxl: pip install pandas openpyxl")
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Для форматування Excel потрібно встановити openpyxl: pip install openpyxl")
    
    ensure_directory_exists(os.path.dirname(file_path) if os.path.dirname(file_path) else '.')
    
    # Створюємо DataFrame з даних
    df = pd.DataFrame(data, columns=fieldnames)
    
    # Зберігаємо в Excel
    df.to_excel(file_path, index=False, engine='openpyxl')
    
    # Застосовуємо форматування
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Стилі для шапки
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Стилі для комірок
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Форматуємо шапку
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx)
        # Використовуємо українську назву, якщо вказано
        if column_headers and fieldname in column_headers:
            cell.value = column_headers[fieldname]
        else:
            cell.value = fieldname
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_style
    
    # Закріплюємо шапку
    ws.freeze_panes = 'A2'
    
    # Знаходимо індекс колонки з посиланнями
    url_column_idx = None
    if 'auction_url' in fieldnames:
        url_column_idx = fieldnames.index('auction_url') + 1
    
    # Стиль для гіперпосилань
    hyperlink_font = Font(underline="single", color="0563C1")
    # Стиль для жирного шрифту
    bold_font = Font(bold=True)
    
    # Форматуємо комірки з даними
    for row_idx in range(2, len(data) + 2):
        row_data = data[row_idx - 2]  # Індекс в масиві data
        # Перевіряємо, чи є додатковий класифікатор 03.07
        has_bold = row_data.get('_has_additional_classification_03_07', False)
        
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = cell_alignment
            
            # Якщо це колонка з посиланнями, створюємо гіперпосилання
            if col_idx == url_column_idx and cell.value:
                url = str(cell.value).strip()
                if url and (url.startswith('http://') or url.startswith('https://')):
                    cell.hyperlink = url
                    cell.value = "Посилання"
                    # Комбінуємо стиль гіперпосилання з жирним, якщо потрібно
                    if has_bold:
                        cell.font = Font(underline="single", color="0563C1", bold=True)
                    else:
                        cell.font = hyperlink_font
                    cell.border = border_style
                elif cell.value and str(cell.value).strip():
                    if has_bold:
                        cell.font = bold_font
                    cell.border = border_style
                else:
                    cell.border = Border()
            # Додаємо сітку тільки якщо є дані
            elif cell.value and str(cell.value).strip():
                if has_bold:
                    cell.font = bold_font
                cell.border = border_style
            else:
                # Прибираємо сітку для порожніх комірок
                cell.border = Border()
    
    # Автоматично підганяємо ширину колонок
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Перевіряємо довжину в шапці
        cell = ws.cell(row=1, column=col_idx)
        if cell.value:
            max_length = len(str(cell.value))
        
        # Перевіряємо довжину в даних
        for row_idx in range(2, len(data) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell_value = str(cell.value)
                # Для багаторядкових значень беремо найдовший рядок
                if '\n' in cell_value:
                    max_line_length = max(len(line) for line in cell_value.split('\n'))
                    max_length = max(max_length, max_line_length)
                else:
                    max_length = max(max_length, len(cell_value))
        
        # Встановлюємо ширину з невеликим запасом
        adjusted_width = min(max_length + 2, 50)  # Максимум 50 символів
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Зберігаємо файл
    wb.save(file_path)
    
    # Підтвердження збереження
    if os.path.exists(file_path):
        file_size = os.path.getsize(file_path)
        print(f"✓ Excel файл збережено: {file_path} ({file_size} байт)")
    else:
        print(f"⚠ Помилка: файл не знайдено після збереження: {file_path}")


def merge_excel_files(file_paths: List[str], output_path: str, fieldnames: List[str], column_headers: Optional[Dict[str, str]] = None) -> None:
    """
    Об'єднує кілька Excel файлів в один.
    
    Args:
        file_paths: Список шляхів до Excel файлів для об'єднання
        output_path: Шлях для збереження об'єднаного файлу
        fieldnames: Список назв колонок (ключі)
        column_headers: Словник з українськими назвами колонок (ключ -> назва)
    """
    if not PANDAS_AVAILABLE:
        raise ImportError("Для об'єднання Excel потрібно встановити pandas та openpyxl: pip install pandas openpyxl")
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("Для форматування Excel потрібно встановити openpyxl: pip install openpyxl")
    
    if not file_paths:
        raise ValueError("Список файлів для об'єднання порожній")
    
    # Читаємо всі файли та об'єднуємо дані
    all_data = []
    for file_path in file_paths:
        if os.path.exists(file_path):
            # Читаємо Excel файл з openpyxl для збереження гіперпосилань
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Знаходимо індекс колонки з посиланнями
            url_column_idx = None
            if 'auction_url' in fieldnames:
                url_column_idx = fieldnames.index('auction_url') + 1
            
            # Читаємо дані з файлу
            rows_data = []
            for row_idx in range(2, ws.max_row + 1):  # Пропускаємо заголовок
                row_data = {}
                row_is_bold = False
                
                # Перевіряємо, чи рядок жирний (перевіряємо першу комірку з даними)
                if len(fieldnames) > 0:
                    first_cell = ws.cell(row=row_idx, column=1)
                    if first_cell.font and first_cell.font.bold:
                        row_is_bold = True
                
                for col_idx, fieldname in enumerate(fieldnames, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    
                    # Якщо це колонка з посиланнями і є гіперпосилання, беремо URL
                    if col_idx == url_column_idx and cell.hyperlink:
                        value = cell.hyperlink.target
                    elif value is None:
                        value = ''
                    
                    row_data[fieldname] = value
                
                # Додаємо службове поле для форматування
                row_data['_has_additional_classification_03_07'] = row_is_bold
                
                # Додаємо рядок тільки якщо він не порожній
                if any(v for v in row_data.values() if v and not (isinstance(v, str) and v.startswith('_'))):
                    rows_data.append(row_data)
            
            all_data.extend(rows_data)
            wb.close()
    
    if not all_data:
        raise ValueError("Немає даних для об'єднання")
    
    # Зберігаємо об'єднані дані (форматування з гіперпосиланнями буде застосовано автоматично)
    save_excel_to_file(all_data, output_path, fieldnames, column_headers)


def find_latest_auction_file(days: int, directory: str = 'archives') -> Optional[str]:
    """
    Знаходить найсвіжіший файл аукціонів за кількістю днів виборки.
    
    Args:
        days: Кількість днів виборки (1 для доби, 7 для тижня)
        directory: Директорія для пошуку (за замовчуванням 'archives')
        
    Returns:
        Шлях до найсвіжішого файлу або None, якщо не знайдено
    """
    import glob
    import re
    
    if not os.path.exists(directory):
        return None
    
    # Шукаємо файли з форматом: prozorro_real_estate_auctions_YYYY-MM-DD_HH-MM-SS[_userID]_daysD.xlsx
    pattern = os.path.join(directory, 'prozorro_real_estate_auctions_*.xlsx')
    files = glob.glob(pattern)
    
    # Фільтруємо файли за кількістю днів
    matching_files = []
    for file_path in files:
        filename = os.path.basename(file_path)
        # Перевіряємо, чи файл містить правильну кількість днів
        # Може бути формат: ..._daysD.xlsx або старий формат без daysD
        match = re.search(rf'_{days}D\.xlsx$', filename)
        if match:
            matching_files.append(file_path)
        elif days == 1:
            # Для старих файлів без daysD в назві, вважаємо їх файлами за добу
            # якщо вони не містять інших днів
            if not re.search(r'_\d+D\.xlsx$', filename):
                matching_files.append(file_path)
    
    if not matching_files:
        return None
    
    # Сортуємо за часом модифікації (найсвіжіший перший)
    matching_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    
    return matching_files[0]


def create_zip_archive(file_path: str, output_path: Optional[str] = None, arcname: Optional[str] = None) -> str:
    """
    Створює ZIP архів з файлу.
    
    Args:
        file_path: Шлях до файлу для архівації
        output_path: Шлях для збереження архіву (опціонально)
        arcname: Назва файлу всередині архіву (опціонально, за замовчуванням використовується basename)
        
    Returns:
        Шлях до створеного ZIP архіву
    """
    import zipfile
    
    if output_path is None:
        output_path = file_path.replace('.xlsx', '.zip')
    
    if arcname is None:
        arcname = os.path.basename(file_path)
    
    ensure_directory_exists(os.path.dirname(output_path) if os.path.dirname(output_path) else '.')
    
    with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(file_path, arcname)
    
    return output_path


def extract_date_range_from_filename(filename: str) -> Optional[tuple]:
    """
    Витягує діапазон дат з назви файлу.
    
    Args:
        filename: Назва файлу у форматі prozorro_real_estate_auctions_YYYY-MM-DD_HH-MM-SS[_userID][_daysD].xlsx
        
    Returns:
        Tuple (date_from, date_to) у форматі datetime або None, якщо не вдалося розпарсити
    """
    import re
    from datetime import timedelta
    
    # Шукаємо дату та час у назві файлу
    match = re.search(r'(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})', filename)
    if not match:
        return None
    
    date_str = match.group(1)
    time_str = match.group(2)
    
    try:
        # Парсимо дату та час
        dt = datetime.strptime(f'{date_str} {time_str}', '%Y-%m-%d %H-%M-%S')
        
        # Шукаємо кількість днів
        days_match = re.search(r'_(\d+)D\.', filename)
        if days_match:
            days = int(days_match.group(1))
            date_to = dt
            date_from = dt - timedelta(days=days)
            return (date_from, date_to)
        else:
            # Для старих файлів без daysD в назві, вважаємо що це файл за добу
            date_to = dt
            date_from = dt - timedelta(days=1)
            return (date_from, date_to)
    except ValueError:
        return None
